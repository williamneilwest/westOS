import json
from datetime import datetime, timezone
from pathlib import Path

from ..tools.document.csv_cleaner import normalize_headers, read_clean_table_rows

from ..models.platform import SessionLocal, SoftwareRegistry, init_platform_db
from .data_source_service import register_source

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None


def _utc_now():
    return datetime.now(timezone.utc)


BOOLEAN_FIELDS = {"phi", "corp_standard", "baa"}
REQUIRED_FIELDS = {"application_name"}


def _normalize_cell(value) -> str:
    return " ".join(str(value or "").strip().split())


def _normalize_bool_like(value) -> str:
    raw = _normalize_cell(value).lower()
    if raw in {"true", "yes", "y", "1"}:
        return "Yes"
    if raw in {"false", "no", "n", "0"}:
        return "No"
    if not raw:
        return ""
    return str(value or "").strip()


def _extract_records_from_rows(headers, raw_rows):
    normalized_headers = normalize_headers(headers or [])
    if not normalized_headers:
        return [], ["File headers were not detected."]

    missing_headers = [field for field in REQUIRED_FIELDS if field not in set(normalized_headers)]
    if missing_headers:
        return [], [f'Missing required headers for: {", ".join(missing_headers)}']

    records = []
    errors = []
    for row_index, row in enumerate(raw_rows, start=2):
        if len(row) != len(normalized_headers):
            errors.append(f"Row {row_index}: column mismatch")
            continue

        normalized = {}
        for index, header in enumerate(normalized_headers):
            cell_value = row[index]
            normalized[header] = _normalize_bool_like(cell_value) if header in BOOLEAN_FIELDS else _normalize_cell(cell_value)

        missing_required = [field for field in REQUIRED_FIELDS if not _normalize_cell(normalized.get(field, ""))]
        if missing_required:
            errors.append(f'Row {row_index}: missing {", ".join(missing_required)}')
            continue
        records.append(normalized)

    return records, errors


def _load_csv_records(source_path: Path):
    with source_path.open("rb") as source_file:
        headers, raw_rows = read_clean_table_rows(source_file, file_type="csv")
    if not headers:
        return [], ["Empty CSV"]
    return _extract_records_from_rows(headers, raw_rows)


def _load_json_records(source_path: Path):
    with source_path.open("rb") as source_file:
        try:
            headers, raw_rows = read_clean_table_rows(source_file, file_type="json")
        except json.JSONDecodeError:
            return [], ["Invalid JSON payload."]

    if not headers:
        return [], ["Empty JSON"]
    return _extract_records_from_rows(headers, raw_rows)


def _load_xlsx_records(source_path: Path):
    if load_workbook is None:
        return [], ["XLSX support is unavailable until openpyxl is installed."]

    workbook = load_workbook(filename=str(source_path), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return [], ["File headers were not detected."]
        headers = [str(cell or "").strip() for cell in header_row]
        raw_rows = []
        for row_values in rows_iter:
            raw_rows.append(
                [
                    "" if index >= len(row_values) or row_values[index] is None else str(row_values[index]).strip()
                    for index in range(len(headers))
                ]
            )
        return _extract_records_from_rows(headers, raw_rows)
    finally:
        workbook.close()


def sync_software_registry_from_latest_file(storage_dir: str | Path):
    """
    Rebuild software registry table from the latest approved source file.
    Returns sync details and validation errors (if any).
    """
    root = Path(storage_dir)
    metadata_path = root / "latest_approved_software.json"
    if not metadata_path.exists():
        return {"synced": False, "reason": "no_metadata"}

    try:
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    except Exception:
        return {"synced": False, "reason": "invalid_metadata"}

    source_path = str(metadata.get("latest_path") or "").strip()
    if not source_path:
        latest_filename = str(metadata.get("latest_filename") or "").strip()
        if latest_filename:
            source_path = str(root / latest_filename)
    if not source_path:
        return {"synced": False, "reason": "missing_source_path"}

    path_obj = Path(source_path)
    if not path_obj.exists():
        latest_filename = str(metadata.get("latest_filename") or "").strip()
        fallback_path = root / latest_filename if latest_filename else None
        if fallback_path is not None and fallback_path.exists():
            path_obj = fallback_path
        else:
            return {"synced": False, "reason": "missing_source_file", "source_path": str(path_obj)}

    extension = path_obj.suffix.lower()
    if extension == ".csv":
        records, errors = _load_csv_records(path_obj)
    elif extension == ".json":
        records, errors = _load_json_records(path_obj)
    elif extension == ".xlsx":
        records, errors = _load_xlsx_records(path_obj)
    else:
        return {"synced": False, "reason": "unsupported_source_type", "source_path": str(path_obj)}

    if not records:
        return {
            "synced": False,
            "reason": "no_valid_records",
            "source_path": str(path_obj),
            "errors": errors[:200],
        }

    save_result = save_software_registry(records, mode="replace")
    return {
        "synced": True,
        "source_path": str(path_obj),
        "records_saved": len(records),
        "errors": errors[:200],
        "total": int(save_result.get("total", 0)),
    }


def save_software_registry(records, mode="replace"):
    """
    Stores cleaned records into source data layer.
    In replace mode, existing rows are removed.
    In upsert mode, rows are merged by vendor_name + application_name.
    """
    init_platform_db()
    session = SessionLocal()
    try:
        now = _utc_now()

        if mode == "replace":
            session.query(SoftwareRegistry).delete(synchronize_session=False)
            for record in records:
                session.add(
                    SoftwareRegistry(
                        vendor_name=record.get("vendor_name", ""),
                        business_function=record.get("business_function", ""),
                        application_name=record.get("application_name", ""),
                        phi=record.get("phi", ""),
                        corp_standard=record.get("corp_standard", ""),
                        baa=record.get("baa", ""),
                        core_level=record.get("core_level", ""),
                        twilight=record.get("twilight", ""),
                        system_owner=record.get("system_owner", ""),
                        hosting_provider=record.get("hosting_provider", ""),
                        deployed_sites=record.get("deployed_sites", ""),
                        description=record.get("description", ""),
                        business_owner=record.get("business_owner", ""),
                        created_at=now,
                        updated_at=now,
                    )
                )
            session.commit()
            register_source(
                key="software_registry",
                name="Software Registry",
                table_name="software_registry",
                row_count=len(records),
            )
            return {"inserted": len(records), "updated": 0, "total": len(records)}

        existing_rows = session.query(SoftwareRegistry).all()
        existing_map = {
            (
                str(row.vendor_name or "").strip().lower(),
                str(row.application_name or "").strip().lower(),
            ): row
            for row in existing_rows
        }

        inserted = 0
        updated = 0
        for record in records:
            key = (
                str(record.get("vendor_name", "")).strip().lower(),
                str(record.get("application_name", "")).strip().lower(),
            )
            current = existing_map.get(key)
            if current is None:
                session.add(
                    SoftwareRegistry(
                        vendor_name=record.get("vendor_name", ""),
                        business_function=record.get("business_function", ""),
                        application_name=record.get("application_name", ""),
                        phi=record.get("phi", ""),
                        corp_standard=record.get("corp_standard", ""),
                        baa=record.get("baa", ""),
                        core_level=record.get("core_level", ""),
                        twilight=record.get("twilight", ""),
                        system_owner=record.get("system_owner", ""),
                        hosting_provider=record.get("hosting_provider", ""),
                        deployed_sites=record.get("deployed_sites", ""),
                        description=record.get("description", ""),
                        business_owner=record.get("business_owner", ""),
                        created_at=now,
                        updated_at=now,
                    )
                )
                inserted += 1
                continue

            current.business_function = record.get("business_function", "")
            current.phi = record.get("phi", "")
            current.corp_standard = record.get("corp_standard", "")
            current.baa = record.get("baa", "")
            current.core_level = record.get("core_level", "")
            current.twilight = record.get("twilight", "")
            current.system_owner = record.get("system_owner", "")
            current.hosting_provider = record.get("hosting_provider", "")
            current.deployed_sites = record.get("deployed_sites", "")
            current.description = record.get("description", "")
            current.business_owner = record.get("business_owner", "")
            current.updated_at = now
            updated += 1

        session.commit()
        total = int(session.query(SoftwareRegistry).count())
        register_source(
            key="software_registry",
            name="Software Registry",
            table_name="software_registry",
            row_count=total,
        )
        return {"inserted": inserted, "updated": updated, "total": total}
    finally:
        session.close()
